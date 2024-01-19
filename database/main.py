from openpyxl import Workbook, load_workbook

from openpyxl.utils import get_column_letter
from  array import *
# Creator
# Amber Binza
# Ian Carlo Perilla
# Micheal Tan
# Creating workbook
# cdata = Workbook()
# Sys = cdata.active
# Sys = "Covid-19 Tracker"
# # cdata.save('Covid -19 Tracker.xlsx')
User = input("Enter UserName: ")
password =  input("Enter Password: ")
if User =="Perilla" and password == "12345":
	cdata = load_workbook('Covid -19 Tracker.xlsx')
	Sys = cdata.active
	counter =100
	nums =[]
	arr = []
	array=[]
	Arrage = []
	# Sys = ["Data num","L_name","F_name","Age","Email","Contact #","Address","Vaccinated"]
	# Sys.append(["1.","Tan", "Michael " , "46", "michael.garcia.tan@Gmail.com", "09812357756" , "63 Chico Street , Barangay Mayondon , Los Baños , Laguna","Y"])
	# Sys.append(["2.","Perilla", "Ian ","19", "cally@gmail.com","09235544571", "82 Accacia Street , Barangay Bagong Buhay , Los Baños , Laguna","Y"])
	# Sys.append(["3.","Binza", "Amber", "19", "ambermae@gmail.com", "09334788120" , "3-B Kalantas Street , Barangay Tabi-tabi , Los Baños , Laguna","N"])
	# Sys.append(["4.","Magdiwang"," Ryan " , "23" ,"magdiwang@gmail.com","09667844510" , "311 Banuyo Street , Barangay Batong Malake , Los Baños , Laguna","Y"])
	# Sys.append(["5.","Ibarra", "Crisostomo"  ,"34" ,"ibarra@gmail.com","09558741147", "1-B Narra  Street , Barangay Sicat , Los Baños , Laguna","Y"])
	# Sys.append(["6.","Clara", "Maria"  , "56", "damaso@gmail.com","09234577851", "4th Avenue , Barangay Silang  , Quezon City","Y"])
	# Sys.append(["7.","Sisa", "Maria Leonora", "43" ,"sisa@gmail.com","09247756147", "23 Molave Street , Barangay Masaya , Los Baños , Laguna","Y"])
	# Sys.append(["8.","Tiago", "Kapitan" ,"55", "tiago@gmail.com","09233147789", "82 Karl Street , Barangay Sven , Imus , Cavite","N"])
	# Sys.append(["9.","Salvi", "Bernardo","34", "salvi@gmail.com","09211476215", "551 Karne Street , Barangay Gutom , Bay , Laguna","N"])
	# Sys.append(["10.","Placido","Penitente", "62", "placido@gmail.com","09332578914", "8th Avenue , Barangay Masaya , Quezon City","N" ])
	# Sys.append(["11.","Gomez", "Paulita" ,"25", "gomez@gmail.com","09546548754", "23 Canola Street , Barangay Tahanin , Valenzuela City","Y"])
	# Sys.append(["12.","Pelaez", "Juanito", "54", "pelaez@gmail.com","09215477451","52 Canola Street , Barangay Tahanin , Valenzuela City","Y"])
	# Sys.append(["13.","Tales", "Kabesa" ,"65", "tales@gmail.com","09621542102", "44-A Y.Q. Street , Barangay Masigasig , Caloocan City","N"])
	# Sys.append(["14.","Penchang", "Hermana", "63", "penchang@gmail.com","09877732549", "144 Compound , Barangay Duyan-Duyan , Quezon City","Y"])
	# Sys.append(["15.","Hermana", "Bali" ,"74", "hermana@gmail.com","09265542123", "Unit 3 , Crystalline Building , Barangay Duyan-Duyan , Quezon City","N"])
	# Sys.append(["16.","Sandoval", "Mario", "45", "sandoval@gmail.com ", "09887541235", "12-H Accacia Street , Bagong Buhay , Los Baños , Laguna","N"])
	# Sys.append(["17.","Tadeo", "Impacto" ,"65", "tadeo@gmail.com","09923142154", "313-A Banuyo Street , Barangay Batong Malake , Los Baños , Laguna","N"])
	# Sys.append(["18.","Vergara", "Janice" , "23", "janigurl@gmail.com","09236654108", "313-B Banuyo Street , Barangay Batong Malake , Los Baños , Laguna","N"])
	# Sys.append(["19.","Ruz", "Marie" ,"29", "marieruz12@gmail.com ", "09322145236", "313-B Banuyo Street , Barangay Batong Malake , Los Baños , Laguna","Y"])
	# Sys.append(["20","Mina", "Yasmine" ,"34", "mina.yas@gmail.com", "09664732154", "Unit 12 , Crystalline Building , Barangay Duyan-Duyan , Quezon City","N" ])
	#
	# cdata.save('Covid -19 Tracker.xlsx')
	print(
		"\n---Covid-19 Tracker Record ---"
		"\nPlease select choices from here:"
		"\n1.Search Name"
		"\n2.Search People based on Age"
		"\n3.Add data"
		"\n4.Delete Data"
		"\n5.Show names in alphabetical order"
		"\n6.Show names by age (least to greatest)"
	)
	userinput = input("Enter Desired number: ")

	menu = int(userinput)

	if menu == 1:
		userin2 = input ( "Enter Surname:" )

		print ( "%-18s %-20s %-18s %-23s %-20s %-70s %s" %
				('\033[34m'"Surname", "First Name", "Age", "Email", "Contact", "Address", "Vaccine Status") )

		for row in range ( 1, counter ):
			for col in range ( 1, 8 ):
				char = chr ( 65 + col )
				age = (Sys['A' + str ( row )].value)

				if (age == userin2):
					f_name = Sys['B' + str ( row )].value
					l_name = Sys['A'+ str ( row )].value
					age = Sys['C' + str ( row )].value
					email = Sys['D' + str ( row )].value
					Contact = Sys['E' + str ( row )].value
					Address = Sys['F' + str ( row )].value
					Vacx = Sys['G' + str ( row )].value

					print('\033[36m'"%-18s %-15s %-18s %-23s %-20s %-75s %-20s"%(l_name, f_name, age, email, Contact, Address, Vacx))
					print ( '\033[36m'"Data Number: " + str ( row ) )
					break



	elif menu ==2:
		userin3 = input ( "Enter Age:" )
		print ( "%-10s %-15s %-15s %-33s %-33s %-65s %s" %
				('\033[34m'"Age", "Surname", "First Name", "Email", "Contact", "Address", "Vaccine Status") )
		for row in range ( 1, counter):
			for col in range ( 1, 8 ):
				char = chr ( 65 + col )
				age= (Sys[char + str ( row )].value)

				if (age == userin3):
					f_name = Sys['B' + str ( row )].value
					l_name = Sys['A' + str ( row )].value
					age = Sys[char + str ( row )].value
					email = Sys['D' + str ( row )].value
					Contact = Sys['E' + str ( row )].value
					Address = Sys['F' + str ( row )].value
					Vacx = Sys['G' + str ( row )].value
					print ('\033[36m' "%-5s %-15s %-15s %-33s %-33s %-70s %s" %
							( age, l_name, f_name, email, Contact, Address, Vacx) )
					break

	if menu ==3:
		nd=[]
		ln = input ('\033[34m' "Enter Last name: " )
		nd.append ( ln )
		fn = input ( '\033[34m'"Enter First name: " )
		nd.append ( fn )
		ageeeee = input ( '\033[34m'"Enter Age: " )
		nd.append ( ageeeee )
		emallll = input ( '\033[34m'"Enter Email: " )
		nd.append ( emallll )
		contact = input ( '\033[34m'"Enter Contact#: " )
		nd.append ( contact )
		addd = input ('\033[34m' "Enter Address: " )
		nd.append ( addd )
		vacy = input ( '\033[34m'"Vaccinated or not use(Y or N) : " )
		nd.append ( vacy )
		Sys.append(nd)
		counter += counter+1

		cdata.save ( 'Covid -19 Tracker.xlsx' )
		print('\033[36m'"Data Added successfully")

	if menu ==5:
		index=0
		print ( "%-28s %-21s %-10s %-30s %-28s %-55s %s" %
				( '\033[34m'"Surname", "First Name","Age", "Email", "Contact", "Address", "Vaccine Status") )
		for row in range (2,counter):
			if 	Sys['A' + str( row )].value !=None:
				Surname = Sys['A' + str( row )].value
				array.append( Surname )
				array.sort()
				arrname =  []
				arrname =array

		while index <=counter:
			for row in range ( 2,counter):
				if arrname[index] == Sys['A' + str ( row )].value and  arrname[index] !=None:

							f_name = Sys['B' + str ( row )].value
							l_name = Sys['A' + str ( row )].value
							age = Sys['C' + str ( row )].value
							email = Sys['D' + str ( row )].value
							Contact = Sys['E' + str ( row )].value
							Address = Sys['F' + str ( row )].value
							Vacx = Sys['G' + str ( row )].value
							print ( '\033[36m'"%-25s %-20s %-5s %-33s %-15s %-75s %s" %
									(l_name, f_name, age, email, Contact, Address, Vacx) )
							index+=1
							row =-row+2
				elif arrname[index] != Sys['A' + str ( row )].value:
								row+=1

	if menu == 4:
		userin3 = input("Enter Datanum:")
		data = int(userin3)
		Sys.delete_rows(data)
		cdata.save ( 'Covid -19 Tracker.xlsx' )

	if menu ==6 :
		for row in range (2,counter):
			if Sys['C' + str ( row )].value != None:
				age = Sys['C'+str(row)].value
				vals =int(age)
				arr.append(vals)
				vals = arr
				nums = vals
	def sort(nums):

		for i in range ( len ( nums ) ):
			min = i
			for j in range ( i, len ( nums ) ):
				if nums[j] < nums[min]:
					min = j

			x = nums[i]
			nums[i] = nums[min]
			nums[min] = x
	sort ( nums )
	indexage =0
	if menu ==6:
		print ( "%-20s %-23s %-21s %-30s %-25s %-70s %s" %
				('\033[34m'"Age", "Surname", "First Name", "Email", "Contact", "Address", "Vaccine Status") )

		while indexage <=counter:

			for row in range ( 2,counter):
				if Sys['C' + str ( row )].value != None:
					Age = int(Sys['C' + str ( row )].value)
					if nums[indexage] == Age  and indexage<=counter:
						f_name = Sys['B' + str ( row )].value
						l_name = Sys['A' + str ( row )].value
						age = Sys['C' + str ( row )].value
						email = Sys['D' + str ( row )].value
						Contact = Sys['E' + str ( row )].value
						Address = Sys['F' + str ( row )].value
						Vacx = Sys['G' + str ( row )].value
						print ( '\033[36m'"%-10s %-28s %-21s %-30s %-25s %-75s %s"%
								(age, l_name, f_name, email, Contact, Address, Vacx) )
						indexage+=1
					if nums[indexage] != Sys['C' + str ( row )].value and indexage<counter:
							row+=1


else:
	print("Wrong UserName or Password")